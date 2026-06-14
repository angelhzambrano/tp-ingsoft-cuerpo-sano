from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.http import HttpResponse
from django.db.models import Q, Sum, Count, F
from django.db.models.functions import TruncDate
from datetime import date, timedelta
from collections import defaultdict
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from asistencia.models import Asistencia
from cobros.models import Cobro
from membresias.models import Membresia
from .forms import FiltroAsistenciasForm, FiltroCobrosForm, FiltroMembresiasVencidasForm


@login_required
@require_http_methods(["GET", "POST"])
def reporte_asistencias(request):
    form = FiltroAsistenciasForm(request.GET or None)
    asistencias = Asistencia.objects.all().select_related('miembro').order_by('-fecha', '-hora')

    fecha_inicio = None
    fecha_fin = None

    if form.is_valid():
        if form.cleaned_data.get('fecha_inicio'):
            fecha_inicio = form.cleaned_data['fecha_inicio']
            asistencias = asistencias.filter(fecha__gte=fecha_inicio)
        if form.cleaned_data.get('fecha_fin'):
            fecha_fin = form.cleaned_data['fecha_fin']
            asistencias = asistencias.filter(fecha__lte=fecha_fin)
        if form.cleaned_data.get('miembro'):
            asistencias = asistencias.filter(miembro=form.cleaned_data['miembro'])
        if form.cleaned_data.get('metodo'):
            asistencias = asistencias.filter(metodo=form.cleaned_data['metodo'])

    # Estadísticas detalladas
    asistencias_list = list(asistencias)

    asistencias_por_dia = defaultdict(int)
    asistencias_por_miembro = defaultdict(int)
    for asistencia in asistencias_list:
        asistencias_por_dia[str(asistencia.fecha)] += 1
        asistencias_por_miembro[str(asistencia.miembro)] += 1

    stats = {
        'total': len(asistencias_list),
        'por_metodo': {
            'BARCODE': len([a for a in asistencias_list if a.metodo == 'BARCODE']),
            'MANUAL': len([a for a in asistencias_list if a.metodo == 'MANUAL']),
        },
        'miembros_unicos': len(set(a.miembro_id for a in asistencias_list)),
        'promedio_diario': round(len(asistencias_list) / max(len(asistencias_por_dia), 1), 1),
        'dias_con_registro': len(asistencias_por_dia),
        'top_miembros': sorted(asistencias_por_miembro.items(), key=lambda x: x[1], reverse=True)[:5],
        'asistencias_por_dia_json': json.dumps({
            'labels': sorted(asistencias_por_dia.keys()),
            'data': [asistencias_por_dia[d] for d in sorted(asistencias_por_dia.keys())]
        })
    }

    # Exportar a Excel si se solicita
    if request.GET.get('export') == 'excel':
        return exportar_asistencias_excel(asistencias_list, fecha_inicio, fecha_fin)

    return render(request, 'reportes/asistencias.html', {
        'form': form,
        'asistencias': asistencias,
        'stats': stats
    })


def exportar_asistencias_excel(asistencias, fecha_inicio, fecha_fin):
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    # Encabezados
    headers = ['Fecha', 'Hora', 'Miembro', 'DNI', 'Tipo Miembro', 'Método']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_num, asistencia in enumerate(asistencias, 2):
        ws.cell(row=row_num, column=1).value = asistencia.fecha.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=2).value = asistencia.hora.strftime('%H:%M:%S')
        ws.cell(row=row_num, column=3).value = str(asistencia.miembro)
        ws.cell(row=row_num, column=4).value = asistencia.miembro.dni
        ws.cell(row=row_num, column=5).value = asistencia.miembro.get_tipo_miembro_display()
        ws.cell(row=row_num, column=6).value = 'Código' if asistencia.metodo == 'BARCODE' else 'Manual'

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10

    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"asistencias_{date.today().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@require_http_methods(["GET", "POST"])
def reporte_cobros(request):
    form = FiltroCobrosForm(request.GET or None)
    cobros = Cobro.objects.all().order_by('-fecha_cobro')

    if form.is_valid():
        if form.cleaned_data.get('fecha_inicio'):
            cobros = cobros.filter(fecha_cobro__gte=form.cleaned_data['fecha_inicio'])
        if form.cleaned_data.get('fecha_fin'):
            cobros = cobros.filter(fecha_cobro__lte=form.cleaned_data['fecha_fin'])
        if form.cleaned_data.get('forma_pago'):
            cobros = cobros.filter(forma_pago=form.cleaned_data['forma_pago'])
        if form.cleaned_data.get('miembro'):
            cobros = cobros.filter(miembro=form.cleaned_data['miembro'])

    stats = {
        'total_cobros': cobros.count(),
        'monto_total': cobros.aggregate(Sum('monto_final'))['monto_final__sum'] or 0,
        'descuento_total': cobros.aggregate(Sum('descuento_monto'))['descuento_monto__sum'] or 0,
        'por_forma': {
            'EFECTIVO': cobros.filter(forma_pago='EFECTIVO').count(),
            'TRANSFERENCIA': cobros.filter(forma_pago='TRANSFERENCIA').count(),
            'TARJETA': cobros.filter(forma_pago='TARJETA').count(),
        },
    }

    return render(request, 'reportes/cobros.html', {
        'form': form,
        'cobros': cobros,
        'stats': stats
    })


@login_required
@require_http_methods(["GET", "POST"])
def membresias_vencidas(request):
    form = FiltroMembresiasVencidasForm(request.GET or None)
    membresias = Membresia.objects.filter(estado='VENCIDA').order_by('-fecha_fin')

    if form.is_valid():
        if form.cleaned_data.get('fecha_vencimiento'):
            membresias = membresias.filter(fecha_fin=form.cleaned_data['fecha_vencimiento'])
        if form.cleaned_data.get('tipo_miembro'):
            membresias = membresias.filter(miembro__tipo_miembro=form.cleaned_data['tipo_miembro'])

    stats = {
        'total_vencidas': membresias.count(),
        'monto_total_no_renovado': membresias.aggregate(
            total=Sum('tipo__precio')
        )['total'] or 0,
        'por_tipo_miembro': dict(
            membresias.values('miembro__tipo_miembro')
            .annotate(count=Count('id'))
            .values_list('miembro__tipo_miembro', 'count')
        ),
        'dias_promedio_sin_renovar': (date.today() - membresias.aggregate(
            fecha_promedio=Sum('fecha_fin')
        )['fecha_promedio']).days if membresias.exists() else 0,
    }

    return render(request, 'reportes/membresias_vencidas.html', {
        'form': form,
        'membresias': membresias,
        'stats': stats
    })
